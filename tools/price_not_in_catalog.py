#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Позиции прайс-листа, которых НЕТ в оцифрованных каталогах (БС-215/БС-230/WP17).

Использует тот же двойной ключ сопоставления «округлённый номер + наименование»,
что и apply_prices.py (см. его про потерю точности номера в исходном .xlsx).
Строка прайса считается «есть в каталоге», только если совпали и номер, и имя;
остальные попадают в выгрузку. Достоверный идентификатор в результате — колонка
«Наименование» (содержит обозначение изделия), а не неточный числовой номер.

Запуск (из корня репозитория):  python3 tools/price_not_in_catalog.py
"""
import sys, os, json, re
HERE=os.path.dirname(os.path.abspath(__file__))
ROOT=os.path.dirname(HERE)
sys.path.insert(0, HERE)
import openpyxl
from apply_prices import norm_name, split_price_name, lossy_key, norm_num, _find_header

import glob
_pl=sorted(glob.glob(os.path.join(ROOT,'Прайс-лист*.xlsx')))
PRICE=_pl[-1] if _pl else os.path.join(ROOT,'Прайс-лист.xlsx')
BUILD=os.path.join(ROOT,'catalog_book','.build')

# --- каталог: множество точных номеров + (lossy_key -> имена) для запасного режима ---
from collections import defaultdict
cat_by_key=defaultdict(set)
cat_numbers=set()
manifest=json.load(open(os.path.join(BUILD,'manifest.json'),encoding='utf-8'))
for eqid in manifest['order']:
    units=json.load(open(os.path.join(BUILD,f'{eqid}.json'),encoding='utf-8'))
    for u in units:
        for pl in u['plates']:
            for r in pl['rows']:
                if r['num'].isdigit():
                    cat_numbers.add(r['num'])
                    k=lossy_key(r['num'])
                    if k is not None:
                        cat_by_key[k].add(norm_name(r['name']))

print(f'Каталог: {len(cat_numbers)} уникальных номеров')

# --- прайс-лист (номер точным текстом; заголовок и колонки определяются сами) ---
wb=openpyxl.load_workbook(PRICE,data_only=True)
ws=wb[wb.sheetnames[0]]
hdr, ci_num, ci_name, ci_price = _find_header(ws)
# ЕИ и Статус — если есть, идут между ценой и после; берём по смещению от номера
ci_unit = ci_price-1 if ci_price-1 not in (ci_num,ci_name) else None
ci_status = ci_price+1

def name_match(pname, cnames):
    if not pname: return False
    for cn in cnames:
        if cn and (pname==cn or pname in cn or cn in pname): return True
    return False

not_in=[]  # строки прайса без совпадения в каталоге
matched=0
for row in ws.iter_rows(min_row=hdr+1,values_only=True):
    if ci_num>=len(row): continue
    num=row[ci_num]; name=row[ci_name] if ci_name<len(row) else None
    price=row[ci_price] if ci_price<len(row) else None
    unit=row[ci_unit] if ci_unit is not None and ci_unit<len(row) else None
    status=row[ci_status] if ci_status<len(row) else None
    if num is None and name is None: continue
    try: price=float(str(price).replace('\xa0','').replace(',','.')) if price not in (None,'') else None
    except ValueError: price=None
    ns=norm_num(num)
    # ТОЧНОЕ совпадение по номеру; иначе запасное по округл.+имя
    pname=split_price_name(name)
    k=lossy_key(ns) if ns else None
    is_in = (ns in cat_numbers) or (k is not None and k in cat_by_key and name_match(pname, cat_by_key[k]))
    if is_in:
        matched+=1
    else:
        num_str=ns   # номер уже точный текст
        # категория достоверности
        nm=str(name or '')
        if re.match(r'\s*(68|173)[.\-]', nm):
            cat='Серия БС (проверить)'   # 68.xx/173.xx — может быть узел не из 3 каталогов ИЛИ промах точности
        else:
            cat='Иное оборуд./расходники' # другой станок, двигатель, компрессор, ГСМ — точно не в этих каталогах
        not_in.append([row[0], num_str, (str(name).strip() if name else ''),
                       (str(unit).strip() if unit else ''),
                       (price if price is not None else ''),
                       (str(status).strip() if status else ''), cat])

# сортировка: по цене убыв. (None в конец)
not_in.sort(key=lambda r: (-(r[4] if isinstance(r[4],(int,float)) else -1)))
print(f'Совпало с каталогом: {matched}')
print(f'НЕТ в каталогах: {len(not_in)}')
from collections import Counter
cc=Counter(r[6] for r in not_in)
for k,v in cc.items(): print(f'   {k}: {v}')

# --- выгрузка в xlsx (номенклатурный номер и цена — текстом/числом аккуратно) ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
out=Workbook(); wsx=out.active; wsx.title='Нет в каталогах'
HEAD=['№ п/п','Номенклатурный номер*','Наименование','Ед. изм.','Цена, руб. без НДС','Статус','Категория']
WIDTH=[8,26,52,10,20,14,24]
fill=PatternFill('solid',fgColor='2F4A63'); hf=Font(color='FFFFFF',bold=True)
thin=Side(style='thin',color='D0D4DA'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
wsx.append(HEAD)
for c in range(1,len(HEAD)+1):
    cell=wsx.cell(1,c); cell.fill=fill; cell.font=hf; cell.border=bd
    cell.alignment=Alignment(vertical='center',wrap_text=True)
for r in not_in:
    wsx.append(r)
for i,w in enumerate(WIDTH,1):
    wsx.column_dimensions[get_column_letter(i)].width=w
# номенклатурный номер как текст, цена как число
for row in wsx.iter_rows(min_row=2,max_row=wsx.max_row,max_col=len(HEAD)):
    for cell in row: cell.border=bd; cell.alignment=Alignment(vertical='top',wrap_text=True)
    row[1].number_format='@'
    if isinstance(row[4].value,(int,float)): row[4].number_format='#,##0.00'
wsx.freeze_panes='A2'
wsx.auto_filter.ref=f'A1:{get_column_letter(len(HEAD))}{wsx.max_row}'
# примечание внизу
note=wsx.max_row+2
wsx.cell(note,1,'* Сопоставление с каталогом — ТОЧНОЕ по номенклатурному номеру '
                '(в актуальном прайс-листе номера хранятся полным текстом). '
                'Каталоги для сравнения: БС-215, БС-230, WP17. '
                '«Серия БС (проверить)» — обозначения 68.xx/173.xx, отсутствующие именно в этих трёх каталогах.')
wsx.cell(note,1).font=Font(italic=True,size=9,color='808080')
out.save(os.path.join(ROOT,'Прайс_нет_в_каталогах.xlsx'))
print('Сохранено: Прайс_нет_в_каталогах.xlsx')
