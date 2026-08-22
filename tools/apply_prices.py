#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
apply_prices.py — сопоставить прайс-лист (.xlsx) с уже собранным каталогом
и вписать цены в кэш .build/<id>.json (перед перегенерацией index.html).

Два режима сопоставления, применяются автоматически:

  1. ТОЧНОЕ по номеру. Если в прайс-листе колонка номера хранится как ТЕКСТ
     (полные 18-значные номера с ведущими нулями — правильная выгрузка), номер
     каталога сравнивается с номером прайса строка-в-строку. Надёжно и без
     угадывания.

  2. ПО ПАРЕ «округлённый номер + наименование» — запасной режим для старых
     выгрузок, где номер хранился как ЧИСЛО и потерял точность на длинных
     номерах (Excel режет после ~15 значащих цифр). Тогда точное сравнение
     невозможно, и цена ставится только если совпали и округлённый номер, и имя.

Для каждой детали сначала пробуется точное совпадение, затем — запасное.

Использование:
    python3 tools/apply_prices.py "Прайс-лист 28.07.2026.xlsx" -o catalog_book
"""
import argparse, json, os, re, sys
from collections import defaultdict
import openpyxl

CODE_PREFIX = re.compile(r'^\s*([\dA-Za-zА-Яа-я]{1,4}(?:[.\-][\dA-Za-zА-Яа-я]{1,4}){1,5})\s+(.+?)\s*$')
NORM = re.compile(r'[^0-9A-ZА-ЯЁ ]+')

def norm_name(s):
    if not s: return ''
    s = str(s).upper().replace('Ё', 'Е')
    s = NORM.sub(' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def split_price_name(raw):
    """'68.10.00.100-01    УСТАНОВКА НАСОСНАЯ' -> 'УСТАНОВКА НАСОСНАЯ'.
    Если явного кода-префикса нет — вся строка и есть название."""
    raw = str(raw or '').replace('\xa0', ' ').strip()
    m = CODE_PREFIX.match(raw)
    if m and re.search(r'\d', m.group(1)):
        return norm_name(m.group(2))
    return norm_name(raw)

def lossy_key(n):
    """Округление до ~12 значащих цифр — приводит номер каталога к той точности,
    что осталась в старом (числовом) прайс-листе, чтобы у них были общие ключи."""
    try:
        return float(f"{float(n):.11e}")
    except (ValueError, TypeError):
        return None

def norm_num(x):
    """Номенклатурный номер как чистая строка цифр (без .0, пробелов, \xa0)."""
    if x is None: return ''
    s = str(x).replace('\xa0', '').strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def _find_header(ws):
    """Строка заголовка (со словом 'Номенклатурный') и индексы колонок номер/имя/цена."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        cells = [str(c) if c is not None else '' for c in row]
        joined = ' '.join(cells)
        if 'Номенклатурный' in joined or ('номер' in joined.lower() and 'Наименование' in joined):
            ci_num = ci_name = ci_price = None
            for j, c in enumerate(cells):
                cl = c.lower()
                if ci_num is None and ('номенклатурный' in cl or 'каталожный' in cl): ci_num = j
                if ci_name is None and 'наименование' in cl: ci_name = j
                if ci_price is None and 'цена' in cl: ci_price = j
            if ci_num is None: ci_num = 1
            if ci_name is None: ci_name = 2
            if ci_price is None: ci_price = 4
            return i, ci_num, ci_name, ci_price
    return 1, 1, 2, 4

def load_price_rows(path):
    """Читает все листы, возвращает список записей с ценой:
    {num_exact, key(lossy), name(norm без кода), price}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        hdr, ci_num, ci_name, ci_price = _find_header(ws)
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            if ci_price >= len(row) or ci_num >= len(row): continue
            num, price = row[ci_num], row[ci_price]
            name = row[ci_name] if ci_name < len(row) else ''
            if num is None or price in (None, ''): continue
            try:
                price = float(str(price).replace('\xa0', '').replace(',', '.'))
            except ValueError:
                continue
            ns = norm_num(num)
            rows.append({'num': ns, 'key': lossy_key(ns), 'name': split_price_name(name), 'price': price})
    return rows

def load_catalog_rows(build_dir, eqid):
    path = os.path.join(build_dir, '.build', f'{eqid}.json')
    if not os.path.exists(path): return None
    units = json.load(open(path, encoding='utf-8'))
    out = []
    for u in units:
        for pl in u['plates']:
            for r in pl['rows']:
                if r['num'] and r['num'].isdigit():
                    out.append({'row': r, 'num': r['num'], 'key': lossy_key(r['num']),
                                'name': norm_name(r['name'])})
    return units, out

def build_indexes(price_rows):
    by_num = {}                      # точный номер -> цена
    by_key = defaultdict(list)       # округлённый номер -> [записи] (запасной режим)
    for pr in price_rows:
        if pr['num']:
            by_num.setdefault(pr['num'], pr['price'])
        if pr['key'] is not None:
            by_key[pr['key']].append(pr)
    return by_num, by_key

def match(price_rows, catalog_rows):
    by_num, by_key = build_indexes(price_rows)
    exact = fuzzy = ambiguous = 0
    examples = []
    for cr in catalog_rows:
        price = None; how = None
        # 1) точное совпадение по номеру
        if cr['num'] in by_num:
            price = by_num[cr['num']]; how = 'номер'
            exact += 1
        else:
            # 2) запасное: округлённый номер + имя
            cands = by_key.get(cr['key'], [])
            if cands:
                names = [c for c in cands if c['name'] == cr['name'] and c['name']]
                pick = names[0] if names else None
                if pick is None and cr['name']:
                    sub = [c for c in cands if c['name'] and (c['name'] in cr['name'] or cr['name'] in c['name'])]
                    if len(set(c['price'] for c in sub)) == 1 and sub:
                        pick = sub[0]
                if pick is not None:
                    price = pick['price']; how = 'номер+имя'; fuzzy += 1
                elif len(cands) > 1:
                    ambiguous += 1
        if price is not None:
            cr['row']['price'] = round(price, 2)
            if len(examples) < 6:
                examples.append((cr['num'], cr['name'], how, round(price, 2)))
        elif 'price' in cr['row']:
            del cr['row']['price']   # убрать устаревшую цену от прошлого прайса
    return exact, fuzzy, ambiguous, examples

def main():
    ap = argparse.ArgumentParser(description=__doc__.split('\n\n')[0])
    ap.add_argument('pricefile')
    ap.add_argument('-o', '--outdir', default='catalog_book')
    a = ap.parse_args()

    price_rows = load_price_rows(a.pricefile)
    print(f'Прайс-лист: {len(price_rows)} строк с ценой')

    manifest_path = os.path.join(a.outdir, '.build', 'manifest.json')
    if not os.path.exists(manifest_path):
        sys.exit(f'нет {manifest_path} — сначала соберите каталог build_catalog.py')
    manifest = json.load(open(manifest_path, encoding='utf-8'))

    tot_exact = tot_fuzzy = tot_rows = 0
    for eqid in manifest['order']:
        loaded = load_catalog_rows(a.outdir, eqid)
        if not loaded: continue
        units, catalog_rows = loaded
        exact, fuzzy, ambiguous, examples = match(price_rows, catalog_rows)
        tot_exact += exact; tot_fuzzy += fuzzy; tot_rows += len(catalog_rows)
        print(f'{eqid}: цена у {exact+fuzzy}/{len(catalog_rows)} позиций '
              f'(точно по номеру: {exact}, по номеру+имени: {fuzzy}, неоднозначных без цены: {ambiguous})')
        for num, cname, how, price in examples[:4]:
            print(f'    {num}  «{cname}» [{how}] -> {price:,.2f} ₽'.replace(',', ' '))
        json.dump(units, open(os.path.join(a.outdir, '.build', f'{eqid}.json'), 'w', encoding='utf-8'),
                  ensure_ascii=False, separators=(',', ':'))

    print(f'\nИТОГО: цена у {tot_exact+tot_fuzzy}/{tot_rows} строк каталога '
          f'(точно: {tot_exact}, по имени: {tot_fuzzy}).')
    print('Кэш .build/*.json обновлён. Пересоберите index.html (build_catalog.py '
          'с теми же PDF — кэш-хит подхватит цены).')

if __name__ == '__main__':
    main()
