#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_numbers_with_prices.py — авторитетная выгрузка всех идентификационных
номеров каталога в Excel, где номера ГАРАНТИРОВАННО хранятся как ТЕКСТ.

Зачем: сами номера каталога извлечены верно (полные, с ведущими нулями —
сверено с PDF дословно). Но если открыть их в Excel как ЧИСЛО, Excel срезает
ведущий ноль и округляет хвост после ~15 значащих цифр («номер не 18-значный /
в конце ноль»). Здесь каждая ячейка номера имеет текстовый формат (@) и явно
строковый тип — Excel их не портит.

Цена подтягивается из прайс-листа по паре «округлённый номер + наименование»
(см. apply_prices.py про потерю точности в самом прайсе). Неоднозначные
совпадения (номер совпал, имя — нет) не проставляются автоматически, а выносятся
на отдельный лист «Цена под вопросом» для ручной сверки.

Запуск:  python3 tools/export_numbers_with_prices.py
Выход:   Каталог_номера_с_ценами.xlsx
"""
import os, re, json
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BUILD=os.path.join(ROOT,'catalog_book','.build')
import glob
_pl=sorted(glob.glob(os.path.join(ROOT,'Прайс-лист*.xlsx')))
PRICE=_pl[-1] if _pl else os.path.join(ROOT,'Прайс-лист.xlsx')

EQLABEL={'bs215':'БС-215','bs230':'БС-230','wp17':'WP17'}

# единый источник правды по сопоставлению цен — apply_prices.py
sys=__import__('sys'); sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from apply_prices import load_price_rows, build_indexes, norm_name as norm, lossy_key as lossy

# ---- каталог: уникальные номера ----
man=json.load(open(os.path.join(BUILD,'manifest.json'),encoding='utf-8'))
uniq={}   # num -> {name, eqs:set, units:set, wp:bool}
order=[]
for eqid in man['order']:
    for u in json.load(open(os.path.join(BUILD,f'{eqid}.json'),encoding='utf-8')):
        for pl in u['plates']:
            for r in pl['rows']:
                n=r['num']
                if not n: continue
                if n not in uniq:
                    uniq[n]={'name':r['name'],'eqs':set(),'units':set()}; order.append(n)
                it=uniq[n]
                if not it['name'] and r['name']: it['name']=r['name']
                it['eqs'].add(EQLABEL.get(eqid,eqid)); it['units'].add(u['name'])

# ---- прайс: точный индекс по номеру + запасной по округлённому номеру ----
_price_rows=load_price_rows(PRICE)
BY_NUM, BY_KEY = build_indexes(_price_rows)   # by_num: точный номер->цена; by_key: округл.->[записи]

def match(num,name):
    # 1) точное совпадение по номеру (новый прайс — номера полным текстом)
    if num in BY_NUM:
        return BY_NUM[num], None, 'ok'
    # 2) запасное: округлённый номер + имя (для старого числового прайса)
    k=lossy(num); cn=norm(name)
    cands=BY_KEY.get(k,[]) if k is not None else []
    if not cands: return None,None,'нет в прайсе'
    ex=[c for c in cands if c['name'] and c['name']==cn]
    if ex: return ex[0]['price'],None,'ok'
    sub=[c for c in cands if c['name'] and (c['name'] in cn or cn in c['name'])]
    if len(set(c['price'] for c in sub))==1 and sub: return sub[0]['price'],None,'ok'
    # неоднозначно: кандидаты для ручной сверки -> (price, name_norm, raw_name)
    return None,[(c['price'],c['name'],c['name']) for c in cands],'под вопросом'

# ---- стили ----
HEAD_FILL=PatternFill('solid',fgColor='2F4A63'); HF=Font(color='FFFFFF',bold=True)
thin=Side(style='thin',color='D5D9DE'); BD=Border(left=thin,right=thin,top=thin,bottom=thin)
TOP=Alignment(vertical='top',wrap_text=True)

def style_header(ws,head,widths):
    ws.append(head)
    for c in range(1,len(head)+1):
        cell=ws.cell(1,c); cell.fill=HEAD_FILL; cell.font=HF; cell.border=BD
        cell.alignment=Alignment(vertical='center',wrap_text=True)
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'

wbx=openpyxl.Workbook()

# ===== Лист 1: Все номера =====
ws=wbx.active; ws.title='Все номера'
HEAD=['Идент. номер','Наименование','Тип номера','Оборудование','Узлы','Цена, руб. без НДС']
style_header(ws,HEAD,[24,42,16,16,40,18])
priced=0; review=[]
for n in order:
    it=uniq[n]
    price,cands,status=match(n,it['name'])
    is_wp = it['eqs']=={'WP17'} or (len(n)<15)
    typ='Weichai (WP17)' if is_wp else ('КМЗ (18-значный)' if len(n)>=17 else 'прочий')
    ws.append([n, it['name'], typ, ', '.join(sorted(it['eqs'])), ', '.join(sorted(it['units'])),
               price if price is not None else ''])
    if price is not None: priced+=1
    if status=='под вопросом':
        review.append((n,it['name'],cands))
# формат: номер — ТЕКСТ, цена — число
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=len(HEAD)):
    for cell in row: cell.border=BD; cell.alignment=TOP
    row[0].number_format='@'; row[0].value=str(row[0].value)   # номер строго текст
    if isinstance(row[5].value,(int,float)): row[5].number_format='#,##0.00'
ws.auto_filter.ref=f'A1:{get_column_letter(len(HEAD))}{ws.max_row}'

# ===== Лист 2: Цена под вопросом (ручная сверка) =====
ws2=wbx.create_sheet('Цена под вопросом')
H2=['Идент. номер','Наименование (каталог)','Наименование (прайс)','Цена-кандидат, руб.']
style_header(ws2,H2,[24,40,40,20])
def tokens(s):
    return set(t for t in norm(s).split() if len(t)>=4)
n_review=0
for n,cname,cands in review:
    ctok=tokens(cname)
    if not ctok: continue
    # оставляем только кандидатов с общим значащим словом (>=4 букв) — отсекаем коллизии
    seen=set(); shown=0
    for p,nm,rn in cands:
        if not (tokens(rn) & ctok): continue
        key=(round(p,2),nm)
        if key in seen: continue
        seen.add(key)
        ws2.append([n, cname, rn, p]); shown+=1
        if shown>=3: break
    if shown: n_review+=1
for row in ws2.iter_rows(min_row=2,max_row=ws2.max_row,max_col=len(H2)):
    for cell in row: cell.border=BD; cell.alignment=TOP
    row[0].number_format='@'; row[0].value=str(row[0].value)
    if isinstance(row[3].value,(int,float)): row[3].number_format='#,##0.00'
ws2.auto_filter.ref=f'A1:{get_column_letter(len(H2))}{ws2.max_row}'

# ===== Лист 3: Пояснение =====
ws3=wbx.create_sheet('Пояснение',0)
notes=[
 'Идентификационные номера каталога БС-215 / БС-230 / WP17',
 '',
 'Все номера в столбце «Идент. номер» хранятся как ТЕКСТ и извлечены из',
 'исходных PDF дословно (проверено). Не меняйте формат ячейки на «Числовой» —',
 'Excel срежет ведущий ноль и округлит длинный номер (после ~15 значащих цифр).',
 '',
 f'Всего уникальных номеров: {len(order)}',
 f'С проставленной ценой: {priced}',
 f'Требуют ручной сверки (лист «Цена под вопросом»): {n_review}',
 '',
 'Цена сопоставляется ТОЧНО по номенклатурному номеру (в актуальном прайс-листе',
 'номера хранятся полным текстом с ведущими нулями). Где точного номера нет —',
 'пробуется запасное совпадение по номеру+наименованию.',
 '',
 'Почему у части номеров нет цены:',
 '  • короткие номера Weichai (WP17) — их нет в прайс-листе КМЗ (это норма);',
 '  • деталь отсутствует в прайс-листе;',
 '  • номер есть, но наименование расходится — такие вынесены на лист',
 '    «Цена под вопросом» для ручной сверки (без авто-подстановки неверной цены).',
]
for i,t in enumerate(notes,1):
    ws3.cell(i,1,t)
    if i==1: ws3.cell(i,1).font=Font(bold=True,size=13)
ws3.column_dimensions['A'].width=95

out=os.path.join(ROOT,'Каталог_номера_с_ценами.xlsx')
wbx.save(out)
print(f'Уникальных номеров: {len(order)} | с ценой: {priced} | под вопросом: {len(review)}')
print('Сохранено:', os.path.basename(out))
